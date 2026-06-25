import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

from loan_calc import LoanSummary
from investment_calc import FDSummary, RDSummary, CISummary

# ── Style helpers ─────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="1F4E79")   # dark blue
ALT_FILL      = PatternFill("solid", start_color="D6E4F0")   # light blue
SUMMARY_FILL  = PatternFill("solid", start_color="2E75B6")   # medium blue
GREEN_FILL    = PatternFill("solid", start_color="E2EFDA")   # light green
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF")

HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT    = Font(name="Arial", bold=True, color="1F4E79", size=13)
SUMMARY_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Arial", size=10)
LABEL_FONT    = Font(name="Arial", bold=True, color="1F4E79", size=10)
INPUT_FONT    = Font(name="Arial", color="0000FF", size=10)   # blue = hardcoded input

THIN = Side(style="thin", color="B0C4DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_FMT = '#,##0.00'
PERCENT_FMT  = '0.00%'
INT_FMT      = '#,##0'


def _header_row(ws, row, cols):
    for col_idx, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _data_row(ws, row, values, alt=False):
    fill = ALT_FILL if alt else WHITE_FILL
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = BODY_FONT
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")
        if isinstance(val, float):
            cell.number_format = CURRENCY_FMT
        if col_idx == 1 and isinstance(val, int):
            cell.number_format = INT_FMT
            cell.alignment = Alignment(horizontal="center")


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _summary_block(ws, start_row, items):
    """Write a 2-column summary block (label | value)."""
    for i, (label, value, fmt) in enumerate(items):
        r = start_row + i
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = LABEL_FONT
        lc.fill = GREEN_FILL
        lc.border = BORDER
        lc.alignment = Alignment(horizontal="left")

        vc = ws.cell(row=r, column=2, value=value)
        vc.font = Font(name="Arial", color="0000FF", size=10, bold=True)
        vc.fill = WHITE_FILL
        vc.border = BORDER
        vc.alignment = Alignment(horizontal="right")
        if fmt:
            vc.number_format = fmt


def _title(ws, row, text, col_span):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    ws.row_dimensions[row].height = 22


# ── Loan Export ───────────────────────────────────────────────────────────────

def export_loan_excel(summary: LoanSummary) -> bytes:
    wb = Workbook()

    # ---- Summary sheet ----
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False

    _title(ws_sum, 1, "🏦  Loan Amortization Summary", 4)
    ws_sum.row_dimensions[2].height = 8

    summary_items = [
        ("Loan Principal",         summary.principal,      CURRENCY_FMT),
        ("Annual Interest Rate",   summary.annual_rate/100, PERCENT_FMT),
        ("Tenure (Months)",        summary.tenure_months,  INT_FMT),
        ("Monthly EMI",            summary.emi,            CURRENCY_FMT),
        ("Total Amount Paid",      summary.total_payment,  CURRENCY_FMT),
        ("Total Interest Paid",    summary.total_interest, CURRENCY_FMT),
        ("Interest to Principal %",
         round(summary.total_interest / summary.principal * 100, 2) if summary.principal else 0,
         '0.00"%"'),
    ]
    _summary_block(ws_sum, 3, summary_items)
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 20

    # ---- Amortization Schedule sheet ----
    ws = wb.create_sheet("Amortization Schedule")
    ws.sheet_view.showGridLines = False

    _title(ws, 1, "📋  Monthly Amortization Schedule", 6)
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 30

    cols = ["Month", "Opening Balance (₹)", "EMI (₹)", "Principal (₹)", "Interest (₹)", "Closing Balance (₹)"]
    _header_row(ws, 3, cols)

    for i, row in enumerate(summary.schedule):
        r = i + 4
        values = [
            row.period,
            row.opening_balance,
            row.emi,
            row.principal,
            row.interest,
            row.closing_balance,
        ]
        _data_row(ws, r, values, alt=(i % 2 == 1))

    # Totals row
    last = len(summary.schedule) + 4
    totals = ["TOTAL", "", summary.total_payment, summary.total_payment - summary.total_interest,
              summary.total_interest, ""]
    for col_idx, val in enumerate(totals, start=1):
        cell = ws.cell(row=last, column=col_idx, value=val)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = SUMMARY_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right" if isinstance(val, float) else "center")
        if isinstance(val, float):
            cell.number_format = CURRENCY_FMT

    _set_col_widths(ws, [10, 22, 18, 18, 18, 22])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── FD Export ─────────────────────────────────────────────────────────────────

def export_fd_excel(summary: FDSummary) -> bytes:
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False
    _title(ws_sum, 1, "🏦  Fixed Deposit Summary", 4)
    ws_sum.row_dimensions[2].height = 8

    summary_items = [
        ("Principal Amount",      summary.principal,       CURRENCY_FMT),
        ("Annual Interest Rate",  summary.annual_rate/100, PERCENT_FMT),
        ("Tenure (Months)",       summary.tenure_months,   INT_FMT),
        ("Compounding",           summary.compounding,     None),
        ("Maturity Amount",       summary.maturity_amount, CURRENCY_FMT),
        ("Total Interest Earned", summary.total_interest,  CURRENCY_FMT),
    ]
    _summary_block(ws_sum, 3, summary_items)
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 20

    ws = wb.create_sheet("Monthly Breakup")
    ws.sheet_view.showGridLines = False
    _title(ws, 1, "📋  FD Monthly Breakup", 5)
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 30

    cols = ["Month", "Opening Balance (₹)", "Interest Earned (₹)", "Closing Balance (₹)", "Cumulative Interest (₹)"]
    _header_row(ws, 3, cols)

    for i, row in enumerate(summary.schedule):
        r = i + 4
        _data_row(ws, r, [row.period, row.opening_balance, row.interest_earned,
                           row.closing_balance, row.cumulative_interest], alt=(i % 2 == 1))

    _set_col_widths(ws, [10, 22, 22, 22, 25])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── RD Export ─────────────────────────────────────────────────────────────────

def export_rd_excel(summary: RDSummary) -> bytes:
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False
    _title(ws_sum, 1, "🏦  Recurring Deposit Summary", 4)
    ws_sum.row_dimensions[2].height = 8

    summary_items = [
        ("Monthly Installment",   summary.monthly_installment, CURRENCY_FMT),
        ("Annual Interest Rate",  summary.annual_rate/100,     PERCENT_FMT),
        ("Tenure (Months)",       summary.tenure_months,       INT_FMT),
        ("Total Deposited",       summary.total_deposited,     CURRENCY_FMT),
        ("Maturity Amount",       summary.maturity_amount,     CURRENCY_FMT),
        ("Total Interest Earned", summary.total_interest,      CURRENCY_FMT),
    ]
    _summary_block(ws_sum, 3, summary_items)
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 20

    ws = wb.create_sheet("Monthly Breakup")
    ws.sheet_view.showGridLines = False
    _title(ws, 1, "📋  RD Monthly Breakup", 7)
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 30

    cols = ["Month", "Installment (₹)", "Opening Balance (₹)", "Interest Earned (₹)",
            "Closing Balance (₹)", "Total Deposited (₹)", "Cumulative Interest (₹)"]
    _header_row(ws, 3, cols)

    for i, row in enumerate(summary.schedule):
        r = i + 4
        _data_row(ws, r, [row.period, row.installment, row.opening_balance,
                           row.interest_earned, row.closing_balance,
                           row.total_deposited, row.cumulative_interest], alt=(i % 2 == 1))

    _set_col_widths(ws, [10, 18, 20, 20, 20, 20, 22])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── CI Export ─────────────────────────────────────────────────────────────────

def export_ci_excel(summary: CISummary) -> bytes:
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False
    _title(ws_sum, 1, "🏦  Compound Interest Investment Summary", 4)
    ws_sum.row_dimensions[2].height = 8

    summary_items = [
        ("Principal Amount",          summary.principal,           CURRENCY_FMT),
        ("Monthly Recurring Amount",  summary.recurring_amount,    CURRENCY_FMT),
        ("Total Contributions",       summary.total_contributions, CURRENCY_FMT),
        ("Annual Interest Rate",      summary.annual_rate/100,     PERCENT_FMT),
        ("Tenure (Years)",            summary.tenure_years,        INT_FMT),
        ("Compounding Frequency",     summary.compounding,         None),
        ("Maturity Amount",           summary.maturity_amount,     CURRENCY_FMT),
        ("Total Interest Earned",     summary.total_interest,      CURRENCY_FMT),
        ("Effective Yield",
         round(summary.total_interest / summary.principal * 100, 2) if summary.principal else 0,
         '0.00"%"'),
    ]
    _summary_block(ws_sum, 3, summary_items)
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 20

    ws = wb.create_sheet("Yearly Breakup")
    ws.sheet_view.showGridLines = False
    _title(ws, 1, "📋  Compound Interest Yearly Breakup", 5)
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 30

    cols = ["Year", "Opening Balance (₹)", "Interest Earned (₹)", "Closing Balance (₹)", "Cumulative Interest (₹)"]
    _header_row(ws, 3, cols)

    for i, row in enumerate(summary.schedule):
        r = i + 4
        _data_row(ws, r, [row.year, row.opening_balance, row.interest_earned,
                           row.closing_balance, row.cumulative_interest], alt=(i % 2 == 1))

    _set_col_widths(ws, [10, 22, 22, 22, 25])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_comparison_excel(summaries) -> bytes:
    """Export multiple loan summaries into a single workbook.

    `summaries` should be an iterable of tuples (name, LoanSummary).
    """
    wb = Workbook()

    # Summary sheet with metrics for all loans
    ws = wb.active
    ws.title = "Comparison Summary"
    ws.sheet_view.showGridLines = False

    _title(ws, 1, "🔁  Loan Comparison Summary", max(5, len(summaries) + 1))
    _header_row(ws, 3, ["Loan", "EMI (₹)", "Tenure (months)", "Total Payment (₹)", "Total Interest (₹)"])

    for i, (name, s) in enumerate(summaries):
        r = 4 + i
        values = [name, s.emi, s.tenure_months, s.total_payment, s.total_interest]
        _data_row(ws, r, values, alt=(i % 2 == 1))

    _set_col_widths(ws, [20, 18, 18, 20, 20])

    # Add individual amortization sheets
    for name, s in summaries:
        ws_s = wb.create_sheet(title=name[:31])
        ws_s.sheet_view.showGridLines = False
        _title(ws_s, 1, f"📋  {name} — Amortization", 6)
        cols = ["Month", "Opening Balance (₹)", "EMI (₹)", "Extra Payment (₹)", "Principal (₹)", "Interest (₹)", "Closing Balance (₹)"]
        _header_row(ws_s, 3, cols)
        for i, row in enumerate(s.schedule):
            r = i + 4
            values = [row.period, row.opening_balance, row.emi, getattr(row, 'extra_payment', 0.0), row.principal, row.interest, row.closing_balance]
            _data_row(ws_s, r, values, alt=(i % 2 == 1))
        _set_col_widths(ws_s, [10, 22, 18, 18, 18, 22])

    # Build a ChartData sheet with Month + each loan outstanding so we can create a native chart
    chart_ws = wb.create_sheet('ChartData')
    # Determine max months
    max_months = max(len(s.schedule) for _, s in summaries)
    # Header row
    headers = ['Month'] + [name for name, _ in summaries]
    _header_row(chart_ws, 1, headers)
    # Fill rows
    for m in range(1, max_months + 1):
        row_vals = [m]
        for _, s in summaries:
            if m <= len(s.schedule):
                row_vals.append(s.schedule[m-1].opening_balance)
            else:
                row_vals.append(None)
        _data_row(chart_ws, 1 + m, row_vals, alt=(m % 2 == 1))
    _set_col_widths(chart_ws, [10] + [18]*len(summaries))

    # Create LineChart referencing the ChartData range
    try:
        chart = LineChart()
        chart.title = "Outstanding Balance Comparison"
        chart.y_axis.title = 'Outstanding Balance'
        chart.x_axis.title = 'Month'
        data_ref = Reference(chart_ws, min_col=2, min_row=1, max_col=1+len(summaries), max_row=1+max_months)
        cats = Reference(chart_ws, min_col=1, min_row=2, max_row=1+max_months)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 20
        chart.height = 7
        ws.add_chart(chart, 'H2')
    except Exception:
        # non-fatal: if chart creation fails, skip
        pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()